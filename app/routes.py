import datetime
import io
import os
from datetime import timedelta, datetime

import jwt
import openpyxl
from flask import *
from openpyxl.styles import Alignment, Side, Border, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

from . import redis
from .config import SECRET_KEY, SQLALCHEMY_DATABASE_URI
from .models import *

new_agro = Blueprint('new_agro', __name__)

success = {'msg': 'ok!'}
wrong = {'msg': 'Something went Wrong!'}
not_exist = {'msg': 'Not Exist in Table!'}


@new_agro.route("/login", methods=['POST'])
def login_a():
    """User Login"""
    username = request.form.get('username')
    password = request.form.get('password')
    try:
        user = User.query.filter_by(username=username).first()
    except:
        return jsonify({'msg': 'chort'})

    if user:
        ch = user.check_password(password)
        if ch:
            token = jwt.encode({
                'public_id': user.id,
                'exp': datetime.datetime.now() + + timedelta(days=100)
            }, SECRET_KEY, algorithm="HS256")
            print(token)
            # user.last_login = datetime.datetime.now()
            # db.session.commit()
            return jsonify({
                'token': token,
                # 'role': [x.format() for x in user.role_metas][0].get("role_name"),
                'msg': "ok",

            }), 200
        return jsonify({
            'msg': "incorrect"
        }), 401
    print('CHORT')
    return jsonify({
        'msg': "not found"
    }), 404


@new_agro.route('/get_mfies')
def getMfies():
    """Get all mfies"""
    mfies = Mfies.query.all()
    return jsonify([x.format() for x in mfies])


@new_agro.route('/get_mfy')
def getMfy():
    """Get Mfy By ID"""
    try:
        id = request.args.get('id')
        mfy = Mfies.query.get(id)
        return jsonify(mfy.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/add_mfy', methods=['POST', 'GET'])
def AddMfy():
    '''
    Add mfie: if GET it will return all regions and districts
    else it will add new mfie
    '''
    if request.method != 'POST':
        regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz, ).all()
        districts = db.session.query(Districts.district_code, Districts.name_ru, Districts.name_uz).all()
        lst = []
        for region in regions:
            lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})
        lst2 = []
        for district in districts:
            lst2.append({'district_code': district[0], 'name_ru': district[1], 'name_uz': district[2]})
        return jsonify(lst, lst2)

    try:
        mfy = Mfies(name_uz=request.form.get('name_uz'), name_ru=request.form.get('name_ru'),
                    region_code=request.form.get('region_code'), district_code=request.form.get('district_code'),
                    mfy_code=request.form.get('mfy_code'))
        mfy.save()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/delete_mfy')
def deleteMfy():
    """It will Delete Mfy from table by ID"""
    try:
        id = request.args.get('id')
        mfy = Mfies.query.get(id)

        mfy.delete()

        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/update_mfy', methods=['POST', 'GET'])
def UpdateMfy():
    if request.method == 'POST':
        try:
            id = request.form.get('id')
            mfy = Mfies.query.get(id)

            mfy.name_ru = request.form.get('name_ru')
            mfy.name_ru = request.form.get('name_uz')
            mfy.region_code = request.form.get('region_code')
            mfy.district_code = request.form.get('district_code')
            mfy.mfy_code = request.form.get('mfy_code')

            db.session.commit()

            return jsonify(success)
        except Exception as e:
            print(e)
            return jsonify(wrong)
    regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz, ).all()
    districts = db.session.query(Districts.district_code, Districts.name_ru, Districts.name_uz).all()
    lst = []
    for region in regions:
        lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})
    lst2 = []
    for district in districts:
        lst2.append({'district_code': district[0], 'name_ru': district[1], 'name_uz': district[2]})
    return jsonify(lst, lst2)


@new_agro.route('/add_region', methods=['POST'])
def addRegion():
    """Add One Region """
    try:
        region = Regions(name_uz=request.form.get('name_uz'), name_ru=request.form.get('name_ru'),
                         region_code=int(request.form.get('region_code')))
        region.save()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/update_region', methods=['POST'])
def UpdateRegion():
    """Update Region by ID"""
    try:
        id = request.form.get('id')
        region = Regions.query.get(id)

        region.name_uz = request.form.get('name_uz')
        region.name_ru = request.form.get('name_ru')
        region.region_code = request.form.get('region_code')

        db.session.commit()
        return jsonify(success)

    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/get_region')
def GetRegion():
    """GET Region by ID"""
    try:
        id = request.args.get('id')
        region = Regions.query.get(id)

        return jsonify(region.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_regions')
def AllRegions():
    """Get All Regions"""
    try:
        prev = None
        next = None
        page = request.args.get('page', 1, type=int)
        regions = Regions.query.all()  # .order_by('id').paginate(page=page, per_page=5)
        total_pages = []
        # for page_num in regions.iter_pages(left_edge=1, right_edge=2, left_current=1, right_current=2):
        #     total_pages.append(page_num)
        # if regions.has_prev:
        #     prev = regions.prev_num
        #
        # if regions.has_next:
        #     next=regions.next_num

        # return jsonify([x.format() for x in regions], [{'total_page': total_pages, 'prev_num': prev, 'next_num':next, 'current_page': page}])
        return jsonify([x.format() for x in regions])
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/delete_region')
def DeletRegion():
    """Delete Region By ID"""
    try:
        id = request.args.get('id')
        region = Regions.query.get(id)

        region.delete()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/add_district', methods=['POST', 'GET'])
def AddDistrict():
    """Add District: if method=POST it will add District to Table
    else it will get all Regions info"""
    if request.method != 'POST':
        try:
            regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz).all()

            lst = []
            for region in regions:
                lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})

            return jsonify(lst)
        except Exception as e:
            print(e)
            return jsonify(wrong)
    try:
        district = Districts(name_uz=request.form.get('name_uz'), name_ru=request.form.get('name_ru'),
                             region_code=request.form.get('region_code'),
                             district_code=request.form.get('district_code'))

        district.save()

        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/update_district', methods=['POST', 'GET'])
def UpdateDistrict():
    """Update District"""
    if request.method == 'POST':
        try:
            id = request.form.get('id')
            district = Districts.query.get(id)

            district.name_uz = request.form.get('name_uz')
            district.name_ru = request.form.get('name_ru')
            district.region_code = request.form.get('region_code')
            district.district_code = request.form.get('district_code')

            db.session.commit()
            return jsonify(success)
        except Exception as e:
            print(e)
            return jsonify(wrong)
    try:
        regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz).all()

        lst = []
        for region in regions:
            lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})

        return jsonify(lst)
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/get_district')
def GetDistrict():
    """GET District by ID"""
    try:
        id = request.args.get('id')
        district = Districts.query.get(id)

        return jsonify(district.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_districts')
def GetAllDistricts():
    """GET ALL DISTRICTS"""
    try:
        districts = Districts.query.all()

        return jsonify([x.format() for x in districts])
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/delete_district')
def DeleteDistrict():
    """Delete District by ID"""
    try:
        id = request.args.get('id')
        district = Districts.query.get(id)

        district.delete()
        return jsonify(success)

    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/create_field', methods=['POST', 'GET'])
def createField():
    if request.method == 'POST':
        try:

            field = Fields(length=request.form.get('length'), area=request.form.get('area'),
                           width=request.form.get('width'),
                           obstacle_degree=request.form.get('obstacle_degree'),
                           stone_degree=request.form.get('stone_degree'),
                           skewness_degree=request.form.get('skewness_degree'), zone=request.form.get('zone'),
                           number=request.form.get('number'),
                           shape_degree=request.form.get('shape_degree'), farm_id=request.form.get('farm_id'))

            field.save()
            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)

    try:
        farms = db.session.query(Farms.id, ).all()

        lst = []

        for farm in farms:
            lst.append([{'farm_id': farm[0]}])

            return jsonify(lst)
    except Exception as e:
        print(e)
        return jsonify(wrong)


@new_agro.route('/add_farm', methods=['POST', 'GET'])
def add_farm():
    if request.method == 'POST':
        try:
            farm = Farms(region_code=request.form.get('region_code'), district_code=request.form.get('district_code'),
                         name=request.form.get('name'), stir=request.form.get('stir'),
                         qfy_name=request.form.get('qfy_name'),
                         street=request.form.get('street'), phone=request.form.get('phone'),
                         fax=request.form.get('fax'),
                         email=request.form.get('email'), postal_code=request.form.get('postal_code'),
                         farmer_name=request.form.get('farmer_name'),
                         accountant_name=request.form.get('accountant_name'),
                         agronomist_name=request.form.get('agronomist_name'),
                         engineer_name=request.form.get('engineer_name'), mfo=request.form.get('mfo'),
                         bank_name=request.form.get('bank_name'),
                         bar_code=request.form.get('bar_code'), mfy_code=request.form.get('mfy_code'))

            farm.save()
            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)

    regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz, ).all()
    districts = db.session.query(Districts.district_code, Districts.name_ru, Districts.name_uz).all()
    lst = []
    for region in regions:
        lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})
    lst2 = []
    for district in districts:
        lst2.append({'district_code': district[0], 'name_ru': district[1], 'name_uz': district[2]})
    return jsonify(lst, lst2)


@new_agro.route('/update_farm', methods=['POST', 'GET'])
def update_farm():
    if request.method == 'POST':
        try:
            id = request.form.get('id')
            farm = Farms.query.get(id)

            farm.region_code = request.form.get('region_code')
            farm.district_code = request.form.get('district_code')
            farm.name = request.form.get('name')
            farm.stir = request.form.get('stir')
            farm.qfy_name = request.form.get('qfy_name')
            farm.street = request.form.get('street')
            farm.phone = request.form.get('phone')
            farm.fax = request.form.get('fax')
            farm.email = request.form.get('email')
            farm.postal_code = request.form.get('postal_code')
            farm.farmer_name = request.form.get('farmer_name')
            farm.accountant_name = request.form.get('accountant_name')
            farm.agronomist_name = request.form.get('agronomist_name')
            farm.engineer_name = request.form.get('engineer_name')
            farm.mfo = request.form.get('mfo')
            farm.bank_name = request.form.get('bank_name')
            farm.bar_code = request.form.get('bar_code')
            farm.mfy_code = request.form.get('mfy_code')

            db.session.commit()

            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)

    regions = db.session.query(Regions.region_code, Regions.name_ru, Regions.name_uz, ).all()
    districts = db.session.query(Districts.district_code, Districts.name_ru, Districts.name_uz).all()
    lst = []
    for region in regions:
        lst.append({'region_code': region[0], 'name_ru': region[1], 'name_uz': region[2]})
    lst2 = []
    for district in districts:
        lst2.append({'district_code': district[0], 'name_ru': district[1], 'name_uz': district[2]})
    return jsonify(lst, lst2)


@new_agro.route('/get_farm')
def get_farm():
    """GET Farm by ID"""
    try:
        id = request.args.get('id')
        farm = Farms.query.get(id)

        return jsonify(farm.format())

    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_farms')
def get_all_farms():
    farms = Farms.query.all()

    return jsonify([x.format() for x in farms])


@new_agro.route('/delete_farm')
def delete_farm():
    try:
        id = request.args.get('id')
        farm = Farms.query.get(id)

        farm.delete()

        return jsonify(success)

    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/add_machine', methods=['POST', 'GET'])
def add_machine():
    """Add Machine"""
    if request.method == 'POST':
        try:
            machine = Machines(type=request.form.get('type'), inventor_number=request.form.get('inventor_number'),
                               model=request.form.get('model'), name=request.form.get('name'),
                               price=request.form.get('price'),
                               balance=request.form.get('balance'),
                               technical_resource=request.form.get('technical_resource'),
                               annual_amortization=request.form.get('annual_amortization'),
                               annual_repair=request.form.get('annual_repair'),
                               annual_service=request.form.get('annual_service'))

            machine.save()
            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)


@new_agro.route('/update_machine', methods=['POST', 'GET'])
def update_machine():
    """Update Machine"""
    if request.method == 'POST':
        try:
            id = request.form.get('id')
            machine = Machines.query.get(id)

            machine.type = request.form.get('type')
            machine.inventor_number = request.form.get('inventor_number')
            machine.model = request.form.get('model')
            machine.name = request.form.get('name')
            machine.price = request.form.get('price')
            machine.balance = request.form.get('balance')
            machine.technical_resource = request.form.get('technical_resource')
            machine.annual_amortization = request.form.get('annual_amortization')
            machine.annual_repair = request.form.get('annual_repair')
            machine.annual_service = request.form.get('annual_service')

            db.session.commit()
            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)


@new_agro.route('/get_machine')
def get_machine():
    """GET Machine"""
    try:
        id = request.args.get('id')
        machine = Machines.query.get(id)

        return jsonify(machine.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_machines')
def get_all_machines():
    """Get All Machines"""
    machines = Machines.query.all()
    return jsonify([x.format() for x in machines])


@new_agro.route('/delete_machine')
def delete_machine():
    """Delete Machine"""
    try:
        id = request.args.get('id')
        machine = Machines.query.get(id)

        machine.delete()

        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/add_plant_type', methods=['POST', 'GET'])
def add_plant_type():
    """Add Plant Type"""
    if request.method == 'POST':
        try:
            plant_type = Plant_Types(name_uz=request.form.get('name_uz'), name_ru=request.form.get('name_ru'),
                                     slug=request.form.get('slug'))

            plant_type.save()
            return jsonify(success)
        except Exception as e:
            print(e)
            return jsonify(wrong)


@new_agro.route('/update_plant_type', methods=['POST'])
def update_plant_type():
    """Update Plant Type"""
    if request.method == 'POST':
        try:
            id = request.form.get('id')
            plant_type = Plant_Types.query.get(id)

            plant_type.name_uz = request.form.get('name_uz')
            plant_type.name_ru = request.form.get('name_ru')
            plant_type.slug = request.form.get('slug')

            db.session.commit()
            return jsonify(success)
        except Exception as e:
            print(e)
            return jsonify(wrong)


@new_agro.route('/get_plant_type')
def get_plant_type():
    """Get Plant Type"""
    try:
        id = request.args.get('id')
        plant_type = Plant_Types.query.get(id)
        return jsonify(plant_type.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_plant_types')
def get_all_plant_types():
    """Get All Plant Types"""
    plant_types = Plant_Types.query.all()
    return jsonify([x.format() for x in plant_types])


@new_agro.route('/delete_plant_type')
def delete_plant_type():
    """Delete Plant Type"""
    try:
        id = request.args.get('id')
        plant_type = Plant_Types.query.get(id)

        plant_type.delete()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)


# @new_agro.route('/add_arrangement')
# def Add_Arrangement():
#     tech_card = Tech_Card(model=request.form.get('model'), machine_name=request.form.get('machine'),
#                               unit=request.form.get('unit'), total=request.form.get('total'),
#                               per_shift=request.form.get('per_shift'), machine_shift=request.form.get('machine_shift'),
#                               employee_shift=request.form.get('employee_shift'), employee=request.form.get('employee'), machine=request.form.get('machine'))
#
#     db.session.add(tech_card)
#     db.session.flush()
#
#     if tech_card.machine != None:
#         machine_shift = tech_card.total / tech_card.per_shift
#
#         tech_card.machine_shift = machine_shift
#         db.session.commit()
#
#
#     elif tech_card.employee != None:
#         machine_shift = tech_card.total / tech_card.per_shift
#
#         tech_card.machine_shift = machine_shift
#         db.session.commit()


@new_agro.route('/add_arrangement', methods=['POST', 'GET'])
def add_arrangement():
    if request.method == 'POST':
        try:
            arrange = Arrangement(name=request.form.get('name'), phase=request.form.get('phase'),
                                  plant_id=request.form.get('plant_id'),
                                  number_of_days=request.form.get('number_of_days'),
                                  unit=request.form.get('unit'), balance_norm=request.form.get('balance_norm'),
                                  discharge=request.form.get('discharge'),
                                  discharge_human=request.form.get('discharge_human'),
                                  square_procent=request.form.get('square_procent'),
                                  gektar_norma=request.form.get('gektar_norma'),

                                  )

            if arrange.gektar_norma == '':
                arrange.gektar_norma = 0

            arrange.square_procent = arrange.square_procent / 100

            db.session.add(arrange)
            db.session.flush()

            # print(arrange.id)

            metaplantarrange = MetaPlantArrange(plant_type_id=arrange.plant_id, arrangement_id=arrange.id,
                                                index=request.form.get('index'))

            db.session.add_all([arrange, metaplantarrange])
            db.session.commit()

            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)

    plant_types = db.session.query(Plant_Types.id, Plant_Types.name_uz, Plant_Types.name_ru).all()

    lst = []
    for plant_type in plant_types:
        lst.append({
            'id': plant_type[0],
            'name_uz': plant_type[1],
            'name_ru': plant_type[2]
        })

    return jsonify(lst)


@new_agro.route('/update_arrangement', methods=['POST', 'GET'])
def update_arrangement():
    if request.method == ['POST']:
        try:
            id = request.form.get('id')
            arrangement = Arrangement.query.get(id)
            metaplantarrange = MetaPlantArrange.query.filter_by(arrangement_id=id).first()

            arrangement.name = request.form.get('name')
            arrangement.phase = request.form.get('phase')
            arrangement.plant_id = request.form.get('plant_id')
            arrangement.number_of_days = request.form.get('number_of_days')
            arrangement.unit = request.form.get('unit')
            arrangement.discharge = request.form.get('discharge')
            arrangement.discharge_human = request.form.get('discharge_human')
            arrangement.square_procent = request.form.get('square_procent')
            arrangement.gektar_norma = request.form.get('gektar_norma')
            arrangement.start_time = request.form.get('start_time')
            arrangement.end_time = request.form.get('end_time')

            metaplantarrange.plant_type_id = request.form.get('plant_id')
            metaplantarrange.index = request.form.get('index')

            db.session.commit()
            return jsonify(success)

        except Exception as e:
            print(e)
            return jsonify(wrong)

    plant_types = db.session.query(Plant_Types.id, Plant_Types.name_uz, Plant_Types.name_ru)

    lst = []
    for plant_type in plant_types:
        lst.append({
            'id': plant_type[0],
            'name_uz': plant_type[1],
            'name_ru': plant_type[2]
        })

    return jsonify(lst)


@new_agro.route('/get_arrangement')
def get_arrangement():
    try:
        id = request.args.get('id')
        arrangement = Arrangement.query.get(id)

        return jsonify(arrangement.format())
    except Exception as e:
        print(e)
        return jsonify(not_exist)


@new_agro.route('/get_all_arrangements')
def get_all_arrangements():
    per_page = request.args.get('per_page')
    page = request.args.get('page', 1, type=int)

    if per_page == None:

        arrangements = Arrangement.query.order_by('id').paginate(page=page, per_page=20)
        total_pages = []
        for page_num in arrangements.iter_pages(left_edge=1, right_edge=2, left_current=1, right_current=2):
            total_pages.append(page_num)

        if arrangements.has_prev:
            prev = arrangements.prev_num
        else:
            prev = None

        if arrangements.has_next:
            next = arrangements.next_num
        else:
            next = None
        return jsonify([x.format() for x in arrangements],
                       [{'total_page': page_num, 'from': prev, 'to': next, 'current_page': page}])

    page = request.args.get('page', 1, type=int)
    arrangements = Arrangement.query.order_by('id').paginate(page=page, per_page=int(per_page))

    total_pages = []
    for page_num in arrangements.iter_pages(left_edge=1, right_edge=2, left_current=1, right_current=2):
        total_pages.append(page_num)
    if arrangements.has_prev:
        prev = arrangements.prev_num
    else:
        prev = None

    if arrangements.has_next:
        next = arrangements.next_num
    else:
        next = None
    return jsonify([x.format() for x in arrangements],
                   [{'total_page': page_num, 'from': prev, 'to': next, 'current_page': page}])


@new_agro.route('/delete_arrangement')
def delete_arrangement():
    try:
        id = request.args.get('id')
        arrangement = Arrangement.query.get(id)

        arrangement.delete()

        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)


import math


@new_agro.route('/execute/techcard', methods=['POST'])
def execute_techcard():
    ekin_id = request.form.get('ekin_type_id')
    gis_area = request.form.get('gis_area')
    # shift_time = 7.0
    shift_coefficient = 1.0

    arrangements = db.session.query(Arrangement).join(MetaPlantArrange,
                                                      MetaPlantArrange.arrangement_id == Arrangement.id).order_by(
        Arrangement.index).filter(
        MetaPlantArrange.plant_type_id == ekin_id).order_by(MetaPlantArrange.index)

    ret_data = []
    mechanizaror = []
    human = []

    for x in arrangements:
        # print(x)
        # varmeta = db.session.query(VarMeta.value).join(MetaPlantArrange,
        #                                                MetaPlantArrange.id == VarMeta.metaplantarrange_id).filter(
        #     MetaPlantArrange.plant_type_id == ekin_id, MetaPlantArrange.arrangement_id == x.id).all()

        # tractor = db.session.query(Tractors.tractor_norm, Tractors.discharge, Tractors.discharge_human, Tractors.name, Tractors.model).filter(
        #     MetaPlantArrange.tractors_id == Tractors.id, MetaPlantArrange.arrangement_id == x.id).first()

        ar_metas = ArrangementMeta.query.order_by(ArrangementMeta.index).all()
        index = db.session.query(MetaPlantArrange.index).filter(x.id == MetaPlantArrange.arrangement_id).first()[0]

        formulas = []
        for i in ar_metas:
            # print(i.value)
            formulas.append(i.value)
        a = x.gektar_norma

        b = x.square_procent if x.square_procent else None
        # print(a, b)

        if formulas[0]:
            formula = formulas[0].replace("c", gis_area) if gis_area else None
            formula = formula.replace("a", str(a))
            formula = formula.replace("b", str(b))
        # print(formula)
        result = round(eval(formula), 2)  # if x.unit != 'га' else float(varmeta[1][0]) * 100

        if x.unit == 'га':
            result = float(x.square_procent) * 100

        if formulas[1]:
            formula2 = formulas[1].replace('tractor', str(float(x.balance_norm)))
            formula2 = formula2.replace('shift_coeficent', str(shift_coefficient))
        workload = eval(formula2)
        if formulas[2]:
            formula3 = formulas[2].replace('overall', str(result))
            formula3 = formula3.replace('day_norma', str(float(x.balance_norm)))
            formula3 = formula3.replace('koef', str(shift_coefficient))

        day_of_shift = eval(formula3)

        if formulas[3]:
            formula3 = formulas[3].replace('day_of_shift', str(float(day_of_shift)))
            formula3 = formula3.replace('number_of_days', str(float(x.number_of_days)))
        jalo = eval(formula3)
        # result = round(eval(formula3)) if x.unit != 'га' else float(varmeta[1][0]) * 100

        if formulas[4]:
            formula4 = formulas[4].replace('shift_time', str(int(x.shift_continuity)))
            # print(shift_time)
            formula4 = formula4.replace('day_of_shift', str((day_of_shift)))
        workhours = eval(formula4)

        gg = {
            'arrangements_name': x.name,
            'phase': x.phase.name,
            "plant_id": x.plant_id,
            'id': x.id,
            'days_to_finish': float(x.number_of_days),
            '1_gektar_norma': float(x.gektar_norma),
            'maydon_procent': float(x.square_procent) * 100,
            'maydon_gektar': float(x.square_procent) * 100,
            "result": result,
            'tractor_norma': float(x.balance_norm),
            'shift_time': x.smenalik_koeffitsiyenti,
            'shift_coefficient': x.shift_continuity,
            # 'workload': float(tractor[0]) * shift_coefficient,
            'workload': workload,
            # 'days_of_shift': round(result / (float(tractor[0]) * shift_coefficient), 2),
            'days_of_shift': round(day_of_shift, 2),
            # 'day_of_shift': round((result / (float(tractor[0]) * shift_coefficient)) / shift_coefficient, 2) if tractor[1] else None,
            'day_of_shift': round(day_of_shift, 2),
            # 'Jalo': round(round((result / (float(tractor[0]) * shift_coefficient)) / shift_coefficient, 2) / int(x.number_of_days)) if tractor[1] else None,
            'Jalo': math.ceil(jalo) if x.discharge else None,
            # 'work_hours': shift_time * round((result / (float(tractor[0]) * shift_coefficient)) / shift_coefficient, 2) if tractor[1] else None,
            'work_hours': round(workhours, 2) if x.discharge else None,
            'discharge': x.discharge,
            'discharge_human': x.discharge_human,
            'start_time': x.start_time,
            'end_time': x.end_time,
            'unit': x.unit,
            'index': index

        }

        ret_data.append(gg)
        mechanizaror.append(round(day_of_shift, 2)) if x.discharge != '' else None
        human.append(round(day_of_shift, 2)) if x.discharge_human != '' else None

    i = {
        'overall_mechanizator': round(sum(mechanizaror), 2),
        'overall_human': round(sum(human), 2)

    }

    return jsonify(ret_data, [i])


import xlrd


@new_agro.route('/get_excel', methods=['POST'])
def get_exel():
    file = request.files.get('file')
    book = xlrd.open_workbook(file_contents=file.read())
    sh = book.sheet_by_index(0)

    for rx in range(1, sh.nrows):

        arrange = Arrangement(name=sh.cell_value(rx, 0), phase_id=sh.cell_value(rx, 10),
                              plant_id=int(sh.cell_value(rx, 7)),
                              number_of_days=sh.cell_value(rx, 13),
                              unit=sh.cell_value(rx, 12), balance_norm=sh.cell_value(rx, 3),
                              discharge=sh.cell_value(rx, 5), discharge_human=sh.cell_value(rx, 6),
                              square_procent=sh.cell_value(rx, 2), gektar_norma=sh.cell_value(rx, 1),
                              row_space=sh.cell_value(rx, 8), smenalik_koeffitsiyenti=sh.cell_value(rx, 9),
                              shift_continuity=sh.cell_value(rx, 4), index=sh.cell_value(rx, 11),
                              bir_birlika=(sh.cell_value(rx, 14)), worker=sh.cell_value(rx, 15),
                              agregat=sh.cell_value(rx, 16), asosiy_mashina_soni=sh.cell_value(rx, 17),
                              kushimcha_mashina_soni=sh.cell_value(rx, 18),
                              birictirilgan_mexanizator=sh.cell_value(rx, 19), group_index=sh.cell_value(rx, 20)
                              )

        if arrange.gektar_norma == '':
            arrange.gektar_norma = 0

        if arrange.row_space == '':
            arrange.row_space = 0

        if arrange.shift_continuity == '':
            arrange.shift_continuity = 0

        if arrange.bir_birlika == '':
            arrange.bir_birlika = 0

        if arrange.asosiy_mashina_soni == '':
            arrange.asosiy_mashina_soni = 0

        if arrange.kushimcha_mashina_soni == '':
            arrange.kushimcha_mashina_soni = 0

        if arrange.birictirilgan_mexanizator == '':
            arrange.birictirilgan_mexanizator = 0

        db.session.add_all([arrange])
        db.session.commit()

    return jsonify(success)


@cached_property
@new_agro.route('/tech_card')
def tech_card() -> Arrangement:
    ekin_id = request.args.get('ekin_type_id')
    gis_area = request.args.get('gis_area')
    cadastor = request.args.get('cadastor')

    cashed = redis.get(f'thech_card:{cadastor}{gis_area}{ekin_id}')

    if cashed:
        print('cashed')
        return json.loads(cashed)

    passport = EarthPasport.query.filter_by(cadastor_num=cadastor, gis_area=gis_area).first()

    if passport:
        #
        er_haydashdan = passport.jadvalfour.er_hayd * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        zanjirli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.zanjir
        boshka_ishlarida = passport.jadvalfour.boshka_ish * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        gildirakli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.gildir

    boshka_ishlarida = 1

    last_id_oil = Oil.query.order_by(Oil.id.desc()).first().id
    oil_price = Oil.query.get(last_id_oil).price

    last_id_tamirlash = Tamirlash_uchun.query.order_by(Tamirlash_uchun.id.desc()).first().id
    tamirlash = Tamirlash_uchun.query.get(last_id_tamirlash)

    last_id_amortizaciya = Amortizatiya_uchun.query.order_by(Amortizatiya_uchun.id.desc()).first().id
    amortizaciya = Amortizatiya_uchun.query.get(last_id_amortizaciya)

    discharges = Discharge.query.all()

    last_id_tuzatish = Tuzatish_coef.query.order_by(Tuzatish_coef.id).first().id
    tuzatish = Tuzatish_coef.query.get(last_id_tuzatish)

    # ar_copy_exist = Arrangement_copy.query.filter_by(cadastor=cadastor, plant_id=ekin_id, gis_area=gis_area).first()

    # if ar_copy_exist:
    #     print('exist')
    #     return tech_card_copy(cadastor, gis_area, ekin_id, tuzatish, oil_price, tamirlash, amortizaciya, discharges,
    #                           district_code )
    over_all_tractor = []

    parameters = [

        {'tuzatish': tuzatish.format(oil_price)},
        {'tamirlash': tamirlash.format()},
        {'amortizaciya': amortizaciya.format()},
        {'discharges': [x.format() for x in discharges]}
    ]

    # if phase_in:
    #     phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).filter(
    #         Arrangement.plant_id == ekin_id, Arrangement.phase_id == phase_in).order_by(Phase.id).all()
    #
    #     for i in phases:
    #         for k in i.arrangements:
    #             ls = k.tech_card(round(gis_area), ekin_id, boshka_ishlarida, oil_price, tamirlash, amortizaciya)
    #             # print(ls)
    #             over_all_tractor.append(ls)
    #
    #     res_tr = []
    #     res_hm = []
    #     for i in over_all_tractor:
    #         # print(i['day_of_shift'])
    #         res_tr.append(i['days_of_shift'])
    #         res_hm.append(i['days_of_shift_human'])
    #     # print(res_tr)
    #
    #     sum_all = {
    #         'mechanizator_sum': round(sum(res_tr), 2),
    #         'human_sum': round(sum(res_hm), 2)
    #     }
    #
    #     return jsonify([sum_all], over_all_tractor)
    # else:
    #     phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).filter(
    #         Arrangement.plant_id == ekin_id)
    # print(phases)
    # ar_l = db.session.query(Arrangement).filter(Arrangement.plant_id == ekin_id, Arrangement.phase_id == phase_in )
    # for i in phases:
    #     for k in i.arrangements:
    #         ls = k.tech_card(gis_area, ekin_id)
    #         print(ls)
    #         over_all_tractor.append(ls)

    # all_phases = Phase.query.order_by(Phase.id).all()

    # all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).filter(
    #     Arrangement.plant_id == ekin_id, ).order_by(Phase.id).all()

    all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).order_by(Phase.id).all()

    ret_data = []
    for x in all_phases:
        # print(x)
        ret_format = {
            'id': x.id,
            'name': x.name,
            'arrangements': [],
            'phase_overall': {
                'days_of_shift': 0,
                "days_of_shift_human": 0,
                "work_machine": 0,
                "work_human": 0,
                "jami_xaj_miga": 0,
                "amortizciya": 0,
                "tamirlash": 0,
                "boshka": 0,
                "jami": 0,
            }
        }
        for i in db.session.query(Arrangement).filter(
                Arrangement.plant_id == ekin_id,
                Arrangement.phase_id == x.id).order_by(Arrangement.index).all():
            # print(i)
            current = i.tech_card(gis_area, ekin_id, oil_price, tamirlash, amortizaciya, boshka_ishlarida)
            # print(type(float(round(current['days_of_shift_human'], 2))))
            ret_format['arrangements'].append(current)

            ret_format['phase_overall']['days_of_shift'] += round(current['days_of_shift'], 0)
            ret_format['phase_overall']['days_of_shift_human'] += round(current['days_of_shift_human'], 0)
            ret_format['phase_overall']['work_machine'] += round(current['work_machine'], 0)
            ret_format['phase_overall']['work_human'] += round(current['work_human'], 0)
            ret_format['phase_overall']['jami_xaj_miga'] += round(current['jami_xaj_miga'], 0)
            ret_format['phase_overall']['amortizciya'] += round(current['amortizciya'], 0)
            ret_format['phase_overall']['tamirlash'] += round(current['tamirlash'], 0)
            ret_format['phase_overall']['boshka'] += round(current['boshka'], 0)
            ret_format['phase_overall']['jami'] += round(current['jami'], 0)

        ret_data.append(ret_format)
        # sts = sorted(ret_format, key=lambda k: k[0])
        # print(ret_data)
    redis.set(f'thech_card:{cadastor}{gis_area}{ekin_id}', json.dumps({'parameters': parameters, "phases": ret_data}),
              ex=600)
    return jsonify({'parameters': parameters, "phases": ret_data})


def add_time(district_code, name, index, cadastor, phase_id, id):
    try:
        arr_dates = Arr_for_set_time.query.filter_by(district_code=district_code, name=name, index=index,
                                                     phase_id=phase_id).first()
        arr_copy = Arrangement_copy.query.filter_by(id=id, name=name, index=index, cadastor=cadastor,
                                                    phase_id=phase_id).first()

        arr_copy.start_time = arr_dates.start_time
        arr_copy.end_time = arr_dates.end_time
        arr_copy.district_code = arr_dates.district_code
        arr_copy.number_of_days = arr_dates.continuity if arr_dates.continuity else 15

        db.session.commit()
        return
    except Exception as e:
        print(e)
        return


@new_agro.route('/tech_card_copy')
def tech_card_copy() -> Arrangement_copy:
    ekin_id = request.args.get('ekin_type_id')
    gis_area = request.args.get('gis_area')
    cadastor = request.args.get('cadastor')
    district_code = request.args.get('district_code')
    contour_number = request.args.get('contour_number')

    cashed = redis.get(f'thech_card_copy:{cadastor}{gis_area}{ekin_id}{contour_number}')

    if cashed:
        print('cashed')
        return json.loads(cashed)

    passport = EarthPasport.query.filter_by(cadastor_num=cadastor, gis_area=gis_area,
                                            counter_number=contour_number).first()

    if passport:
        er_haydashdan = passport.jadvalfour.er_hayd * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        zanjirli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.zanjir
        boshka_ishlarida = passport.jadvalfour.boshka_ish * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        gildirakli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.gildir

        tusiclar_proc = (int(passport.bir) * 4 + int(passport.asosi) * 20 + int(passport.ayrisimon) * 78 + int(
            passport.turt) * 176) / (float(gis_area) * 100)
        print(tusiclar_proc)
        passport_obj = {
            'er_haydashdan': er_haydashdan,
            'zanjirli_tractor_bulsa': zanjirli_tractor_bulsa,
            'boshka_ishlarida': boshka_ishlarida,
            'gildirakli_tractor_bulsa': gildirakli_tractor_bulsa,
            'tusiclar_proc': round(tusiclar_proc, 2)

        }
    else:

        er_haydashdan = 1
        zanjirli_tractor_bulsa = 1
        boshka_ishlarida = 1
        gildirakli_tractor_bulsa = 1

        passport_obj = {
            'er_haydashdan': er_haydashdan,
            'zanjirli_tractor_bulsa': zanjirli_tractor_bulsa,
            'boshka_ishlarida': boshka_ishlarida,
            'gildirakli_tractor_bulsa': gildirakli_tractor_bulsa,

        }

    all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).join(Arrangement_copy,
                                                                                                  Arrangement_copy.arrangement_id == Arrangement.id).filter(
        Arrangement.id == Arrangement_copy.arrangement_id).order_by(Phase.id).all()

    last_id_oil = Oil.query.order_by(Oil.id.desc()).first().id
    oil_price = Oil.query.get(last_id_oil).price

    last_id_tamirlash = Tamirlash_uchun.query.order_by(Tamirlash_uchun.id.desc()).first().id
    tamirlash = Tamirlash_uchun.query.get(last_id_tamirlash)

    last_id_amortizaciya = Amortizatiya_uchun.query.order_by(Amortizatiya_uchun.id.desc()).first().id
    amortizaciya = Amortizatiya_uchun.query.get(last_id_amortizaciya)

    discharges = Discharge.query.all()

    last_id_tuzatish = Tuzatish_coef.query.order_by(Tuzatish_coef.id).first().id
    tuzatish = Tuzatish_coef.query.get(last_id_tuzatish)

    parameters = [

        {'tuzatish': tuzatish.format(oil_price)},
        {'tamirlash': tamirlash.format()},
        {'amortizaciya': amortizaciya.format()},
        {'discharges': [x.format() for x in discharges]}
    ]

    ret_data = []
    for x in all_phases:

        ret_format = {
            'id': x.id,
            'name': x.name,
            'arrangements': [],
            'phase_overall': {
                'days_of_shift': 0,
                "days_of_shift_human": 0,
                "work_machine": 0,
                "work_human": 0,
                "jami_xaj_miga": 0,
                "amortizciya": 0,
                "tamirlash": 0,
                "boshka": 0,
                "jami": 0,
            }
        }

        for i in db.session.query(Arrangement_copy).filter(
                Arrangement_copy.cadastor == cadastor, Arrangement_copy.phase_id == x.id,
                Arrangement_copy.gis_area == gis_area, Arrangement_copy.plant_id == ekin_id,
                Arrangement_copy.contour_number == contour_number).order_by(
            Arrangement_copy.index).all():

            if i.end_time == None and i.start_time == None:
                add_time(district_code, i.name, i.index, cadastor, i.phase_id, i.id)

            current = i.tech_card(i.gis_area, i.plant_id, tuzatish, oil_price, tamirlash, amortizaciya, er_haydashdan,
                                  zanjirli_tractor_bulsa, boshka_ishlarida, gildirakli_tractor_bulsa)

            ret_format['arrangements'].append(current)

            ret_format['phase_overall']['days_of_shift'] += round(current['days_of_shift'], 0)
            ret_format['phase_overall']['days_of_shift_human'] += round(current['days_of_shift_human'], 0)
            ret_format['phase_overall']['work_machine'] += round(current['work_machine'], 0)
            ret_format['phase_overall']['work_human'] += round(current['work_human'], 0)
            ret_format['phase_overall']['jami_xaj_miga'] += round(current['jami_xaj_miga'], 0)
            ret_format['phase_overall']['amortizciya'] += round(current['amortizciya'], 0)
            ret_format['phase_overall']['tamirlash'] += round(current['tamirlash'], 0)
            ret_format['phase_overall']['boshka'] += round(current['boshka'], 0)
            ret_format['phase_overall']['jami'] += round(current['jami'], 0)

        ret_data.append(ret_format)
    #
    # first_phase_start_time = ret_data[0]['arrangements'][0]['start_time']
    # first_phase_end_time = ret_data[1]['arrangements'][0]['start_time']
    # other_phase_start_time = ret_data[1]['arrangements'][0]['start_time']
    # other_phase_end_time = ret_data[-1]['arrangements'][-1]['end_time']
    # first_phase_name = ret_data[0]['name']

    # savedTime = SavedArrTime(cadastor_number=cadastor, gis_area=gis_area, contour_number=contour_number,
    #                          district_code=district_code)
    #
    # db.session.add(savedTime)
    # db.session.flush()

    # print(first_phase_name)
    # print(first_phase_start_time)
    # print(first_phase_end_time)
    # print(other_phase_end_time)
    #
    # if first_phase_start_time == '':
    #     first_phase_start_time = None
    #
    # if first_phase_end_time == '':
    #     first_phase_end_time = None
    #
    # if other_phase_start_time == '':
    #     other_phase_start_time = None
    #
    # if other_phase_end_time == '':
    #     other_phase_end_time = None

    # savedTime.first_phase_start_time = first_phase_start_time
    # savedTime.first_phase_end_time = first_phase_end_time
    # savedTime.other_phase_start_time = other_phase_start_time
    # savedTime.other_phase_end_time = other_phase_end_time

    # db.session.commit()

    redis.set(f'thech_card_copy:{cadastor}{gis_area}{ekin_id}{contour_number}',
              json.dumps({'parameters': parameters, "phases": ret_data, 'passport_obg': passport_obj}), ex=120)
    return jsonify({'parameters': parameters, "phases": ret_data, 'passport_obg': passport_obj})


def get_times(ekin_id, gis_area, cadastor, contour_number):
    try:
        savedTime = SavedArrTime.query.filter_by(cadastor_number=cadastor, gis_area=gis_area,
                                                 contour_number=contour_number, ).first()
        # print(savedTime)

        all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).join(Arrangement_copy,
                                                                                                      Arrangement_copy.arrangement_id == Arrangement.id).filter(
            Arrangement.id == Arrangement_copy.arrangement_id).order_by(Phase.id).all()

        last_id_oil = Oil.query.order_by(Oil.id.desc()).first().id
        oil_price = Oil.query.get(last_id_oil).price

        last_id_tamirlash = Tamirlash_uchun.query.order_by(Tamirlash_uchun.id.desc()).first().id
        tamirlash = Tamirlash_uchun.query.get(last_id_tamirlash)

        last_id_amortizaciya = Amortizatiya_uchun.query.order_by(Amortizatiya_uchun.id.desc()).first().id
        amortizaciya = Amortizatiya_uchun.query.get(last_id_amortizaciya)

        discharges = Discharge.query.all()

        last_id_tuzatish = Tuzatish_coef.query.order_by(Tuzatish_coef.id).first().id
        tuzatish = Tuzatish_coef.query.get(last_id_tuzatish)

        parameters = [

            {'tuzatish': tuzatish.format(oil_price)},
            {'tamirlash': tamirlash.format()},
            {'amortizaciya': amortizaciya.format()},
            {'discharges': [x.format() for x in discharges]}
        ]

        ret_data = []
        for x in all_phases:

            ret_format = {
                'id': x.id,
                'name': x.name,
                'arrangements': [],
                'phase_overall': {
                    'days_of_shift': 0,
                    "days_of_shift_human": 0,
                    "work_machine": 0,
                    "work_human": 0,
                    "jami_xaj_miga": 0,
                    "amortizciya": 0,
                    "tamirlash": 0,
                    "boshka": 0,
                    "jami": 0,
                }
            }

            for i in db.session.query(Arrangement_copy).filter(
                    Arrangement_copy.cadastor == cadastor, Arrangement_copy.phase_id == x.id,
                    Arrangement_copy.gis_area == gis_area, Arrangement_copy.plant_id == ekin_id,
                    Arrangement_copy.contour_number == contour_number).order_by(
                Arrangement_copy.index).all():
                current = i.tech_card(i.gis_area, i.plant_id, tuzatish, oil_price, tamirlash, amortizaciya)
                # print(type(float(round(current['days_of_shift_human'], 2))))
                ret_format['arrangements'].append(current)

                ret_format['phase_overall']['days_of_shift'] += round(current['days_of_shift'], 0)
                ret_format['phase_overall']['days_of_shift_human'] += round(current['days_of_shift_human'], 0)
                ret_format['phase_overall']['work_machine'] += round(current['work_machine'], 0)
                ret_format['phase_overall']['work_human'] += round(current['work_human'], 0)
                ret_format['phase_overall']['jami_xaj_miga'] += round(current['jami_xaj_miga'], 0)
                ret_format['phase_overall']['amortizciya'] += round(current['amortizciya'], 0)
                ret_format['phase_overall']['tamirlash'] += round(current['tamirlash'], 0)
                ret_format['phase_overall']['boshka'] += round(current['boshka'], 0)
                ret_format['phase_overall']['jami'] += round(current['jami'], 0)

            ret_data.append(ret_format)

        first_phase_start_time = ret_data[0]['arrangements'][0]['start_time']
        first_phase_end_time = ret_data[1]['arrangements'][0]['start_time']
        other_phase_start_time = ret_data[1]['arrangements'][0]['start_time']
        other_phase_end_time = ret_data[-1]['arrangements'][-1]['end_time']
        first_phase_name = ret_data[0]['name']

        if savedTime:
            print('exist')
        else:

            # print(first_phase_start_time)
            # print(first_phase_end_time)
            # print(other_phase_end_time)

            if first_phase_start_time == '':
                first_phase_start_time = None

            if first_phase_end_time == '':
                first_phase_end_time = None

            if other_phase_start_time == '':
                other_phase_start_time = None

            if other_phase_end_time == '':
                other_phase_end_time = None

            times = SavedArrTime(first_phase_start_time=first_phase_start_time, first_phase_end_time=first_phase_end_time,
                                 contour_number=contour_number, plant_id=ekin_id,
                                 gis_area=gis_area, other_phase_start_time=other_phase_start_time,
                                 other_phase_end_time=other_phase_end_time, cadastor_number=cadastor,
                                 first_phase_name=first_phase_name)

            db.session.add(times)
            db.session.commit()

        return jsonify({'parameters': parameters, "phases": ret_data})
    except Exception as e:
        print(e)
        return jsonify({'msg': 'error'})


@new_agro.route('/save_tech_card', methods=['POST', 'GET'])
def save_tech_card():
    if request.method == 'POST':
        gis_area = request.form.get('gis_area')
        ekin_id = request.form.get('ekin_type_id')
        cadastor = request.form.get('cadastor')
        contour_number = request.form.get('contour_number')

        # print(contour_number)

        ximicat = request.form.get('ximicat')
        qator_oraliqi = request.form.get('qator_oraliqi')
        nasos = request.form.get('nasos')
        terim = request.form.get('terim')
        mintaqa = request.form.get('mintaqa')
        qoshqator = request.form.get('qoshqator')
        ekish_sharoiti = request.form.get('ekish_sharoiti')
        # print(bool(ximicat))

        # print(shablon.id)

        # shablon = Shablon.query.get(1)

        # print(shablon.arrangements[0].arrangement.name)

        print(type(ximicat))

        follder = Folder.query.filter_by(cad_number=cadastor).first()

        if follder:
            exist_savedcard = SavedTechCard.query.filter_by(cad_number=cadastor, plant_id=ekin_id, gis_area=gis_area,
                                                            folder_id=follder.id, contour_number=contour_number).first()
            if exist_savedcard:
                return jsonify({'msg': 'exist',
                                'ok': False})
            else:
                saved_tech_card = SavedTechCard(cad_number=cadastor, plant_id=ekin_id, gis_area=gis_area,
                                                contour_number=contour_number,
                                                folder_id=follder.id)
                db.session.add(saved_tech_card)
                db.session.commit()
        else:
            print(bool(int(ximicat)))
            print(qator_oraliqi)
            print(nasos)
            print(terim)
            print(mintaqa)
            print(qoshqator)
            print(ekish_sharoiti)
            shablon = Shablon.query.filter_by(ximicat=bool(int(ximicat)), qator_oraliqi=qator_oraliqi, nasos=nasos,
                                              terim=terim,
                                              mintaqa=mintaqa, qoshqator=qoshqator,
                                              ekish_sharoiti=ekish_sharoiti).first()

            # arrrangements = Arrangement.query.filter_by(plant_id=ekin_id).order_by(Arrangement.id)
            arrrangements = sorted(shablon.arrangements, key=lambda shablon: shablon.id)

            for i in arrrangements:
                if request.method == 'POST':
                    arrangementCopy_colls = list_of_cols_from(Arrangement_copy)

                    arrangementCopy = Arrangement_copy()

                    for col in arrangementCopy_colls:
                        val = getattr(i.arrangement, col, None)
                        # val = getattr(i, col, None)

                        setattr(arrangementCopy, col, val)
                        arrangementCopy.cadastor = cadastor
                        arrangementCopy.arrangement_id = i.arrangement.id
                        # arrangementCopy.arrangement_id = i.id
                        arrangementCopy.gis_area = gis_area
                        arrangementCopy.id = None
                        arrangementCopy.plant_id = ekin_id
                        arrangementCopy.contour_number = contour_number

                    db.session.add(arrangementCopy)
                    db.session.commit()

            exist_arr_copy = Arrangement_copy.query.filter_by(gis_area=gis_area, cadastor=cadastor, plant_id=ekin_id,
                                                              contour_number=contour_number)

            for i in exist_arr_copy:
                if i.end_time == None and i.start_time == None:
                    add_time(request.form.get('district_code'), i.name, i.index, cadastor, i.phase_id, i.id)

            get_times(ekin_id, gis_area, cadastor, contour_number)

            folder = Folder(cad_number=cadastor)

            db.session.add(folder)
            db.session.flush()

            saved_tech_card = SavedTechCard(cad_number=cadastor, plant_id=ekin_id, gis_area=gis_area,
                                            folder_id=folder.id, contour_number=contour_number)

            db.session.add_all([saved_tech_card, folder])

            # arr_copy = Arrangement_copy(start_time=i.start_time, end_time=i.end_time, name=i.name,
            #                             number_of_days=i.number_of_days,
            #                             index=i.index, plant_id=i.plant_id, unit=i.unit,
            #                             balance_norm=i.balance_norm,
            #                             discharge=i.discharge,
            #                             discharge_human=i.discharge_human, gektar_norma=i.gektar_norma,
            #                             square_procent=i.square_procent,
            #                             row_space=i.row_space, shift_continuity=i.shift_continuity, phase_id=i.phase_id,
            #                             smenalik_koeffitsiyenti=i.smenalik_koeffitsiyenti, worker=i.worker,
            #                             agregat=i.agregat, bir_birlika=i.bir_birlika, spec_formula=i.spec_formula,
            #                             cadastor=cadastor, gis_area=gis_area, arrangement_id=i.id,
            #                             asosiy_mashina_soni=i.asosiy_mashina_soni,
            #                             kushimcha_mashina_soni=i.kushimcha_mashina_soni,
            #                             birictirilgan_mexanizator=i.birictirilgan_mexanizator,
            #                             group_index=i.group_index)
            # db.session.add(arr_copy)
            # db.session.commit()
            db.session.commit()
    return jsonify({'ok': True})


@new_agro.route('/delete_arr_copy')
def delete_arrangement_copy():
    try:
        id = request.args.get('id')
        arrangement = Arrangement_copy.query.get(id)

        cashed = redis.get(
            f'thech_card_copy:{arrangement.cadastor}{arrangement.gis_area}{arrangement.plant_id}{arrangement.contour_number}')

        arrangement.delete()
        if cashed:
            print('cashed')
            redis.delete(
                f'thech_card_copy:{arrangement.cadastor}{arrangement.gis_area}{arrangement.plant_id}{arrangement.contour_number}')

            return jsonify(success)
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify(not_exist)

    # return jsonify(over_all_tractor)


@new_agro.route('/edit_arr_copy', methods=['POST'])
def edit_arr_copy():
    id = request.form.get('id')
    # start_time = request.form.get('start_time')
    # end_time = request.form.get('end_time')
    # if start_time == '':
    #     start_time = None
    # if end_time == '':
    #     end_time = None
    arr_copy = Arrangement_copy.query.get(id)

    user_col_names = list_of_cols_from(Arrangement_copy)

    cashed = redis.get(
        f'thech_card_copy:{arr_copy.cadastor}{arr_copy.gis_area}{arr_copy.plant_id}{arr_copy.contour_number}')

    try:
        for col in user_col_names:

            val = request.form.get(col)
            if val:
                if val == request.form.get('start_time'):
                    if val == '':
                        val = None
                        setattr(arr_copy, col, val)
                    setattr(arr_copy, col, val)

                if val == request.form.get('end_time'):
                    if val == '':
                        val = None
                        setattr(arr_copy, col, val)
                    setattr(arr_copy, col, val)
                if val == request.form.get('spec_formula'):
                    if val == '':
                        val = None
                        setattr(arr_copy, col, val)
                    setattr(arr_copy, col, val)
                setattr(arr_copy, col, val)

        db.session.commit()

        arr_copies = Arrangement_copy.query.filter(Arrangement_copy.cadastor == arr_copy.cadastor,
                                                   Arrangement_copy.plant_id == arr_copy.plant_id,
                                                   Arrangement_copy.id != arr_copy.id,
                                                   Arrangement_copy.phase_id == arr_copy.phase_id).all()

        for i in arr_copies:
            index = int(i.index)
            if int(arr_copy.index) < index or int(arr_copy.index) == index:
                index += 1
                i.index = index
            db.session.commit()
            # if int(arr_copy.index) >= index:
            #     print(index)
            #     index -= 1
            #     i.index = index

        if cashed:
            print('cashed')
            redis.delete(
                f'thech_card_copy:{arr_copy.cadastor}{arr_copy.gis_area}{arr_copy.plant_id}{arr_copy.contour_number}')
            return jsonify(success)
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445
    # arr_copy.start_time = start_time
    # arr_copy.end_time = end_time
    # arr_copy.name = request.form.get('name')
    # arr_copy.number_of_days = request.form.get('number_of_days')
    # arr_copy.index = request.form.get('index')
    # arr_copy.plant_id = request.form.get('plant_id')
    # arr_copy.unit = request.form.get('unit')
    # arr_copy.balance_norm = request.form.get('balance_norm')
    # arr_copy.discharge = request.form.get('discharge')
    # arr_copy.gektar_norma = request.form.get('gektar_norma')
    # arr_copy.square_procent = request.form.get('square_procent')
    # arr_copy.row_space = request.form.get('row_space')
    # arr_copy.shift_continuity = request.form.get('shift_continuity')
    # arr_copy.phase_id = request.form.get('phase_id')
    # arr_copy.smenalik_koeffitsiyenti = request.form.get('smenalik_koeffitsiyenti')
    # arr_copy.worker = request.form.get('worker')
    # arr_copy.agregat = request.form.get('agregat')
    # arr_copy.bir_birlika = request.form.get('bir_birlika')
    # arr_copy.spec_formula = request.form.get('spec_formula')
    # arr_copy.cadastor = request.form.get('cadastor')
    # arr_copy.gis_area = request.form.get('gis_area')
    # arr_copy.group_index = request.form.get('group_index')
    #
    # db.session.commit()

    return jsonify(success)


# @new_agro.route('/setattr', methods=['POST'])
# def setAttr():
#     ik_cols = list_of_cols_from(Arrangement_copy)
#     print(ik_cols)
#     ik = Arrangement_copy()
#     try:
#         for col in ik_cols:
#             val = request.form.get(col)
#             if val:
#                 setattr(ik, col, val)
#         db.session.add(ik)
#         db.session.commit()
#         return jsonify(success)
#     except Exception as e:
#         print(e)
#         return jsonify({"error in col ": str(e)}), 445

@new_agro.route('/add_copy', methods=['POST', 'GET'])
def add_copy():
    if request.method == 'POST':
        arr_copy_colls = list_of_cols_from(Arrangement_copy)

        arr_copy = Arrangement_copy()

        for col in arr_copy_colls:
            val = request.form.get(col)
            if val == request.form.get('start_time'):
                if val == '':
                    val = None
                    setattr(arr_copy, col, val)
                setattr(arr_copy, col, val)

            if val == request.form.get('end_time'):
                if val == '':
                    val = None
                    setattr(arr_copy, col, val)
                setattr(arr_copy, col, val)
            setattr(arr_copy, col, val)
        cashed = redis.get(
            f'thech_card_copy:{arr_copy.cadastor}{arr_copy.gis_area}{arr_copy.plant_id}{arr_copy.contour_number}')
        print(f'thech_card_copy:{arr_copy.cadastor}{arr_copy.gis_area}{arr_copy.plant_id}{arr_copy.contour_number}')
        db.session.add(arr_copy)
        db.session.commit()

        arr_copies = Arrangement_copy.query.filter(Arrangement_copy.cadastor == str(request.form.get('cadastor')),
                                                   Arrangement_copy.plant_id == int(request.form.get('plant_id')),
                                                   Arrangement_copy.id != arr_copy.id,
                                                   Arrangement_copy.phase_id == int(request.form.get('phase_id'))).all()

        for i in arr_copies:
            index = int(i.index)
            if int(arr_copy.index) <= index:
                index += 1
                i.index = index

            db.session.commit()
        if cashed:
            print('cashed')
            redis.delete(
                f'thech_card_copy:{arr_copy.cadastor}{arr_copy.gis_area}{arr_copy.plant_id}{arr_copy.contour_number}')
            return jsonify(success)
        return jsonify(success)

        # try:
        # start_time = request.form.get('start_time')
        # end_time = request.form.get('end_time')
        # if start_time == '':
        #     start_time = None
        # if end_time == '':
        #     end_time = None
        #
        # arr_copy = Arrangement_copy(start_time=start_time, end_time=end_time,
        #                             name=request.form.get('name'),
        #                             number_of_days=request.form.get('number_of_days'),
        #                             index=request.form.get('index'), plant_id=request.form.get('plant_id'),
        #                             unit=request.form.get('unit'), balance_norm=request.form.get('balance_norm'),
        #                             discharge=request.form.get('discharge'),
        #                             discharge_human=request.form.get('discharge_human'),
        #                             gektar_norma=request.form.get('gektar_norma'),
        #                             square_procent=request.form.get('square_procent'),
        #                             row_space=request.form.get('row_space'),
        #                             shift_continuity=request.form.get('shift_continiyty'),
        #                             phase_id=request.form.get('phase_id'),
        #                             smenalik_koeffitsiyenti=request.form.get('smenalik_koeffitsiyenti'),
        #                             worker=request.form.get('worker'),
        #                             agregat=request.form.get('agregat'), bir_birlika=request.form.get('bir_birlika'),
        #                             spec_formula=request.form.get('spec_formula'),
        #                             cadastor=request.form.get('cadastor'), gis_area=request.form.get('gis_area'),
        #                             arrangement_id=request.form.get('arrangement_id'))
        # db.session.add(arr_copy)
        # db.session.commit()
        #
        # arr_copies = Arrangement_copy.query.filter(Arrangement_copy.cadastor == str(request.form.get('cadastor')),
        #                                            Arrangement_copy.plant_id == int(request.form.get('plant_id')),
        #                                            Arrangement_copy.id != arr_copy.id,
        #                                            Arrangement_copy.phase_id == int(request.form.get('phase_id'))).all()
        #
        # for i in arr_copies:
        #     index = int(i.index)
        #     if int(arr_copy.index) <= index:
        #         index += 1
        #         i.index = index
        #
        #     db.session.commit()
        #
        # return jsonify(success)
        # except Exception as e:
        #     print(e)
        #     return jsonify({'msg': str(e)})



    else:
        plant_type_id = request.args.get('ekin_type_id')
        arrangements = Arrangement.query.filter_by(plant_id=plant_type_id).order_by(Arrangement.index).all()
        ls = []
        for i in arrangements:
            data = {
                'name': i.name,
                'phase': i.phase.name,
                'plant_type_name': i.arrangement_for_plant.name_uz,
                'id': i.id,
                'created_at': i.created_at,
                'start_time': i.start_time.strftime("%d-%m-%Y") if i.start_time else '',
                'end_time': i.end_time.strftime("%d-%m-%Y") if i.end_time else '',
                'updated_at': i.updated_at,
                'name': i.name,
                'number_of_days': i.number_of_days,
                'index': i.index,
                'plant_id': i.plant_id,
                'unit': i.unit,
                'balance_norm': i.balance_norm,
                'discharge': i.discharge,
                'discharge_human': i.discharge_human,
                'gektar_norma': i.gektar_norma,
                'row_space': i.row_space,
                'shift_continuity': i.shift_continuity,
                'phase_id': i.phase_id,
                'smenalik_koeffitsiyenti': i.smenalik_koeffitsiyenti,
                'worker': i.worker,
                'agregat': i.agregat,
                'bir_birlika': i.bir_birlika,
                'spec_formula': i.spec_formula,
                'square_procent': i.square_procent,
                'group_index': i.group_index
            }
            ls.append(data)
        return jsonify(ls)


@new_agro.route('/phases')
def get_all_phases():
    phases = Phase.query.all()
    return [x.format() for x in phases]


@new_agro.route('/get_copy')
def getCopy():
    id = request.args.get('id')
    arr_copy = Arrangement_copy.query.get(id)
    data = {
        'name': arr_copy.name,
        'phase': arr_copy.arrangement_copy.phase.name,
        'plant_type_name': arr_copy.arrangement_for_plant.name_uz,
        'id': arr_copy.id,
        'created_at': arr_copy.created_at,
        'start_time': arr_copy.start_time,
        'end_time': arr_copy.end_time,
        'updated_at': arr_copy.updated_at,
        'arrangements_name': arr_copy.name,
        'number_of_days': arr_copy.number_of_days,
        'index': arr_copy.index,
        'plant_id': arr_copy.plant_id,
        'unit': arr_copy.unit,
        'balance_norm': arr_copy.balance_norm,
        'discharge': arr_copy.discharge,
        'discharge_human': arr_copy.discharge_human,
        'gektar_norma': arr_copy.gektar_norma,
        'square_procent': arr_copy.square_procent,
        'row_space': arr_copy.row_space,
        'shift_continuity': arr_copy.shift_continuity,
        'phase_id': arr_copy.phase_id,
        'smenalik_koeffitsiyenti': arr_copy.smenalik_koeffitsiyenti,
        'worker': arr_copy.worker,
        'agregat': arr_copy.agregat,
        'bir_birlika': arr_copy.bir_birlika,
        'spec_formula': arr_copy.spec_formula,
        'group_index': arr_copy.group_index,

        'square_procent': arr_copy.square_procent * 100

    }

    return jsonify(data)


@new_agro.route('/delete_all_copies')
def delete_all_copies():
    cadastor = request.args.get('cadastor')
    ekin_type_id = request.args.get('ekin_type_id')
    contour_number = request.args.get('contour_number')
    gis_area = request.args.get('gis_area')

    cashed = redis.get(f'thech_card_copy:{cadastor}{gis_area}{ekin_type_id}{contour_number}')
    print(f'thech_card_copy:{cadastor}{gis_area}{ekin_type_id}{contour_number}')

    arr_copies = Arrangement_copy.query.filter_by(cadastor=cadastor, plant_id=ekin_type_id).all()

    for i in arr_copies:
        db.session.delete(i)
        db.session.commit()

    if cashed:
        redis.delete(f'thech_card_copy:{cadastor}{gis_area}{ekin_type_id}{contour_number}')
        return jsonify(success)
    return jsonify(success)


def list_of_cols_from(cls):
    return [column.key for column in cls.__table__.c]


@new_agro.route('/setattr', methods=['POST'])
def setAttr():
    ik_cols = list_of_cols_from(Arrangement_copy)
    # print(ik_cols)
    ik = Arrangement_copy()
    try:
        for col in ik_cols:
            val = request.form.get(col)
            if val:
                setattr(ik, col, val)
        db.session.add(ik)
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.route('/editattr', methods=['POST'])
def edit_idoraviy_user():
    id = request.form.get("id")
    table = Arrangement_copy.query.get(id)

    print("proshel")
    user_col_names = list_of_cols_from(Arrangement_copy)
    try:
        for col in user_col_names:
            # if col in ["active", "daraja", "lavozim", "pnfl", "home_phone", "idoraviy_korxona_id", "role"]:
            #     continue
            val = request.form.get(col)
            if val:
                setattr(table, col, val)
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.route('/jadvalone')
def jadvalone():
    jdone = db.session.query(JadvalOne).filter(JadvalOne.procent == '5 % gacha').first()

    return str(jdone.jadval_one_tuzat.tuzatish_koef_kiymati_smenalik.gidravlika_va_zanjirli)


#
# @new_agro.route('/create_passport', methods=['POST', 'GET'])
# def create_passport():
#     if request.method == 'GET':
#         jadvalone = JadvalOne.query.all()
#         jadvalthree = JadvalThree.query.all()
#
#         # jadvalfour = Jadval_Four.query.all()
#         #
#         # er_kiyilyagilar = ErRelefningKiyaligi.query.all()
#
#         # group_by = db.session.query(Ish_Unumiga.value).group_by(Ish_Unumiga.value)
#         # print(group_by)
#         # for x in group_by:
#         #     print(x[0])
#
#         json_file = {
#             'jadval1': [x.format() for x in jadvalone],
#             'jadval2': [x.format() for x in er_kiyilyagilar],
#             'jadval3': [x.format() for x in jadvalthree],
#             'jadval4': [x.format() for x in jadvalfour],
#         }
#
#         return jsonify(json_file)
#     if request.method == 'POST':
#         passport_colls = list_of_cols_from(EarthPasport)
#
#         passport = EarthPasport()
#
#         for col in passport_colls:
#             val = request.form.get(col)
#
#             setattr(passport, col, val)
#         db.session.add(passport)
#         db.session.commit()
#
#         return jsonify(success)
#
#
# @new_agro.route('/delete_by_group', methods=['POST'])
# def delete_by_group():
#     phase_id = request.form.get('phase_id')
#     group_index = request.form.get('group_index')
#     cadastor = request.form.get('cadastor')
#     plant_id = request.form.get('plant_id')
#     contour_number = request.form.get('contour_number')
#     gis_area = request.form.get('gis_area')
#
#     arrangements = Arrangement_copy.query.filter_by(phase_id=phase_id, group_index=group_index, cadastor=cadastor,
#                                                     plant_id=plant_id, contour_number=contour_number,
#                                                     gis_area=gis_area).all()
#
#
#     cashed = redis.get(f'thech_card_copy:{cadastor}{gis_area}{plant_id}{contour_number}')
#
#     if request.method == 'GET':
#         return jsonify([x.format() for x in arrangements])
#
#     for i in arrangements:
#         db.session.delete(i)
#         db.session.commit()
#     if cashed:
#         print('cashed')
#         redis.delete(f'thech_card_copy:{i.cadastor}{i.gis_area}{i.plant_id}{i.contour_number}')
#         return jsonify(success)
#     return jsonify(success)


@new_agro.post('/update_passport')
def updatePassport():
    id = request.form.get('id')
    passport = EarthPasport.query.get(id)

    user_col_names = list_of_cols_from(EarthPasport)
    # try:

    for col in user_col_names:

        val = request.form.get(col)
        if val:
            setattr(passport, col, val)
    db.session.commit()
    return jsonify(passport.format())


@new_agro.get('/delete_passport')
def DeletePassport():
    id = request.args.get('id')
    passport = EarthPasport.query.get(id)

    db.session.delete(passport)
    db.session.commit()
    return jsonify(success)


@new_agro.route('/districts_by_code', methods=['POST', 'GET'])
def DistrictsByCode():
    region = request.args.get('code')

    district = Districts.query.filter_by(region_code=int(region)).all()

    ls = []
    for i in district:
        ls.append({'name_uz': i.name_uz,
                   'district_code': i.district_code})

    return jsonify(ls)


@new_agro.route('/generate_datetimes', methods=['POST', 'GET'])
def GenerateDAtes():
    if request.method == 'POST':
        try:

            district_code = request.form.get('district_code')
            plant_id = request.form.get('plant_id')

            arr_date = db.session.query(Arr_for_set_time).filter(
                Arr_for_set_time.plant_id == plant_id,
                Arr_for_set_time.district_code == district_code).order_by(Arr_for_set_time.index).first()

            if arr_date:
                # print(arr_date)
                print('exist')
                params = {'district_code': district_code,
                          'plant_id': plant_id}
                # request_url = requests.get(url=f'http://localhost:5001/generate_datetimes', params=params)
                # return request_url.json()
                return 'exist'

            arrrangements = Arrangement.query.filter_by(plant_id=plant_id).order_by(Arrangement.id)

            for i in arrrangements:
                arr_date = Arr_for_set_time(name=i.name, start_time=i.start_time, end_time=i.end_time, index=i.index,
                                            district_code=district_code, plant_id=i.plant_id, phase_id=i.phase_id,
                                            continuity=i.number_of_days)
                db.session.add(arr_date)
                db.session.commit()

            return jsonify(success)
        except Exception as e:
            print(e)
            return jsonify({'msg': 'exist!'}), 445

    district_code = request.args.get('district_code')
    plant_id = request.args.get('plant_id')

    cashed = redis.get(f"generated_dates: {district_code}-{plant_id}")

    # if cashed:
    #     print('cashed')
    #     return json.loads(cashed)

    all_phases = db.session.query(Phase).join(Arr_for_set_time, Arr_for_set_time.phase_id == Phase.id).filter(
        Arr_for_set_time.plant_id == plant_id, ).order_by(Phase.id).all()

    ret_data = []

    for x in all_phases:
        ret_format = {
            'id': x.id,
            'name': x.name,
            'arrangements': [],

        }

        for i in Arr_for_set_time.query.filter_by(plant_id=plant_id, district_code=district_code,
                                                  phase_id=x.id).order_by(
            Arr_for_set_time.index):
            ret_format['arrangements'].append(i.format())

        ret_data.append(ret_format)
        redis.set(f"generated_dates: {district_code}-{plant_id}", json.dumps({"phases": ret_data}), ex=1200)

    return jsonify({"phases": ret_data})


@new_agro.route('/set_datetime', methods=['POST'])
def SetDAte():
    id = request.form.get('id')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    continuity = request.form.get('continuity')

    if start_time == '':
        start_time = None
    if end_time == '':
        end_time = None

    arr_date = Arr_for_set_time.query.get(id)

    arr_date.start_time = start_time
    arr_date.end_time = end_time
    arr_date.continuity = continuity

    db.session.commit()

    return jsonify(success)


import pandas as pd


@new_agro.post('/exl_to_sql')
def EXLtoSql():
    file = request.files.get('file')
    df = pd.read_excel(file)
    df2 = pd.read_excel(file)
    df.pop('model')
    df.to_sql(df2['model'][0], create_engine(SQLALCHEMY_DATABASE_URI), if_exists='append', index=False)
    return jsonify(success)


# @new_agro.get('/pizdec')
# def pizdec():
#     import json
#
#     from app import db
#     from app.models import Plan
#
#     with open('ekin.geojson', 'r', encoding='utf-8') as f:
#         text = json.load(f)
#         # print(text)
#         for i in text['features']:
#             print(i['properties']['globalid'])
#             plan = Plan(
#                 geometry=i['geometry'], globalid=i['properties']['globalid'], region=i['properties']['region'],
#                 district=i['properties']['district'], crop_name=i['properties']['crop_name'],
#                 crop_area=i['properties']['crop_area'], kontur_raqami=i['properties']['kontur_raqami'],
#                 massiv=i['properties']['Massiv'], shape_length=i['properties']['Shape_Length'],
#                 shape_area=i['properties']['Shape_Area'],
#             )
#             db.session.add(plan)
#             db.session.commit()
#         print('done')
#         return 'done'

@new_agro.get('/geojson')
def getGeoJson():
    kontur_num = request.args.get('contour_num')
    plan = Plan.query.filter_by(kontur_raqami=kontur_num).all()

    return jsonify([i.format() for i in plan])


@new_agro.get('/all_folders')
def AllFolders():
    folders = Folder.query.order_by(Folder.id.desc())

    # return jsonify([x.folder_list() for x in folders])
    return list(map(lambda x: x.folder_list(), folders))


@new_agro.get('/inside_folder')
def insideFolder():
    cad_number = request.args.get('cad_number')
    contour_number = request.args.get('contour_number')
    gis_area = request.args.get('gis_area')
    folder = SavedTechCard.query.filter_by(cad_number=cad_number, contour_number=contour_number, gis_area=gis_area)
    return jsonify(list(map(lambda x: x.format(), folder)))


@new_agro.route('/search')
def search():
    keyword = request.args.get('q')
    search_cad = Folder.query.msearch(keyword, fields=['cad_number'])  # limit=6

    # return jsonify([x.folder_list() for x in search_cad])
    return list(map(lambda x: x.folder_list(), search_cad))


@new_agro.get('/check_tech_card')
def checkTechCard():
    cad_number = request.args.get('cad_number')
    contour_number = request.args.get('contour_number')
    arr_copy = Arrangement_copy.query.filter_by(cadastor=cad_number, contour_number=contour_number).first()

    if arr_copy:
        return jsonify({
            'plant_id': arr_copy.plant_id,
            'row_space': arr_copy.row_space
        })
    return jsonify({})


@new_agro.post('/create_exl')
def createExl():
    cadastor = request.form.get('cadastor')
    plant_id = request.form.get('plant_id')
    contour_number = request.form.get('contour_number')
    gis_area = request.form.get('gis_area')

    output = io.BytesIO()

    full_path = os.path.join(current_app.root_path, 'saved_exl', f'{gis_area}-{plant_id}')

    if not os.path.exists(full_path):
        os.makedirs(full_path)

    file_path = os.path.join(full_path, f'{"arrangement"}-{2}.xlsx')

    pd.read_sql(
        # f"SELECT * FROM arrangement_copy WHERE arrangement_copy.cadastor='{17:05:000088258}' AND arrangement_copy.plant_id='{3}' AND arrangement_copy.contour_number='{1156 / 315}' AND arrangement_copy.gis_area='{5.437631366024745}' ORDER BY arrangement_copy.index",
        f"SELECT * FROM arrangement",
        create_engine(SQLALCHEMY_DATABASE_URI)).to_excel(file_path)

    # return Response(output, mimetype="application/ms-excel",
    #                 headers={"Content-Disposition": "attachment;filename=test.xlsx"})

    return file_path


@new_agro.route('/get_exl')
def getUsers():
    ekin_id = request.args.get('ekin_type_id')
    gis_area = request.args.get('gis_area')
    # phase_in = request.form.get('phase')
    cadastor = request.args.get('cadastor')
    district_code = request.args.get('district_code')
    contour_number = request.args.get('contour_number')

    cashed = redis.get(f'exl_tech_card:{cadastor}{gis_area}{ekin_id}')

    if cashed:
        print('cashed')
        ret_data = json.loads(cashed)
    else:
        last_id_oil = Oil.query.order_by(Oil.id.desc()).first().id
        oil_price = Oil.query.get(last_id_oil).price

        last_id_tamirlash = Tamirlash_uchun.query.order_by(Tamirlash_uchun.id.desc()).first().id
        tamirlash = Tamirlash_uchun.query.get(last_id_tamirlash)

        last_id_amortizaciya = Amortizatiya_uchun.query.order_by(Amortizatiya_uchun.id.desc()).first().id
        amortizaciya = Amortizatiya_uchun.query.get(last_id_amortizaciya)

        last_id_tuzatish = Tuzatish_coef.query.order_by(Tuzatish_coef.id).first().id
        tuzatish = Tuzatish_coef.query.get(last_id_tuzatish)

        all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).filter(
            Arrangement.plant_id == ekin_id, ).order_by(Phase.id).all()
        ret_data = []
        for x in all_phases:

            ret_format = {
                'id': x.id,
                'name': x.name,
                'arrangements': [],
                'phase_overall': {
                    'days_of_shift': 0,
                    "days_of_shift_human": 0,
                    "work_machine": 0,
                    "work_human": 0,
                    "jami_xaj_miga": 0,
                    "amortizciya": 0,
                    "tamirlash": 0,
                    "boshka": 0,
                    "jami": 0,
                }
            }

            for i in db.session.query(Arrangement_copy).filter(
                    Arrangement_copy.cadastor == cadastor, Arrangement_copy.phase_id == x.id,
                    Arrangement_copy.gis_area == gis_area, Arrangement_copy.plant_id == ekin_id,
                    Arrangement_copy.contour_number == contour_number).order_by(
                Arrangement_copy.index).all():

                if i.end_time == None and i.start_time == None:
                    add_time(district_code, i.name, i.index, cadastor, i.phase_id, i.id)

                current = i.tech_card_exel(i.gis_area, i.plant_id, tuzatish, oil_price, tamirlash, amortizaciya)
                # print(type(float(round(current['days_of_shift_human'], 2))))
                ret_format['arrangements'].append(current)
                #
                ret_format['phase_overall']['days_of_shift'] += round(current['h'], 0)
                ret_format['phase_overall']['days_of_shift_human'] += round(current['i'], 0)
                ret_format['phase_overall']['work_machine'] += round(current['l'], 0)
                ret_format['phase_overall']['work_human'] += round(current['m'], 0)
                ret_format['phase_overall']['jami_xaj_miga'] += round(current['o'], 0)
                ret_format['phase_overall']['amortizciya'] += round(current['p'], 0)
                ret_format['phase_overall']['tamirlash'] += round(current['q'], 0)
                ret_format['phase_overall']['boshka'] += round(current['r'], 0)
                ret_format['phase_overall']['jami'] += round(current['s'], 0)

            ret_data.append(ret_format)
            # print(ret_data)
            redis.set(f'exl_tech_card:{cadastor}{gis_area}{ekin_id}', json.dumps(ret_data),
                      ex=600)

    output = io.BytesIO()

    workbook = openpyxl.Workbook()

    workbook.remove(workbook.active)

    sheet = workbook.create_sheet('card')

    # ws = workbook.active
    # sheet.column_dimensions['A'].width = 50

    thick = Side(border_style="thin", color="000000")
    # double = Side(border_style="double", color="ff0000")
    #
    sheet.insert_rows(0)
    sheet['A1'].value = '№'

    sheet.merge_cells('A1:A2')
    sheet['A3'].value = 1
    sheet['A1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['A2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['A3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["A3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('B1:B2')
    sheet['B1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['B2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['B3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet.column_dimensions['B'].width = 30
    sheet['B3'].value = 2
    sheet["B3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['B1'].value = 'Агротехник тадбирларнинг номи'
    sheet["B1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["B1"].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('C1:D1')
    sheet['C1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['C2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['C3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['C1'].value = 'Агрегат таркиби'
    sheet["C1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['C2'].value = 'трактор русуми'
    sheet["C2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['C3'].value = 3
    sheet["C3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['D1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['D2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['D3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['D2'].value = 'машина русуми ёки иш воситаси'
    sheet["D2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['D3'].value = 4
    sheet["D3"].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('E1:G1')
    # sheet.column_dimensions['E'].width = 15
    sheet['E1'].value = 'Бажариладиган иш ҳажми'
    sheet["E1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['E2'].value = 'ўлчов бир-лиги'
    sheet['E3'].value = 5
    sheet["E3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["E2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['F2'].value = 'жами миқ-дори'
    sheet['F3'].value = 6
    sheet["F3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet["F2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['E1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['E2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['E3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['F1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['F2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['F3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['G2'].value = 'смена-лик меъёри (ҳисоб)'
    sheet['G3'].value = 7
    sheet["G2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet["G3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['G1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['G2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['G3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    sheet.merge_cells('H1:I1')
    sheet['H1'].value = 'Киши- смена'
    sheet["H1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['H2'].value = 'механи-затор-ники'
    sheet["H2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['H3'].value = 8
    sheet["H3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['I2'].value = 'ишчи-ники'
    sheet["I2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['I3'].value = 9
    sheet["I3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['H1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['H2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['H3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['I1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['I2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['I3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    sheet.merge_cells('J1:K1')
    sheet['J1'].value = 'Тариф разряди'
    sheet["J1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['J1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['J2'].value = 'механи-затор-ники'
    sheet["J2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['J2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['J3'].value = 10
    sheet["J3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['J3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['K2'].value = 'ишчи-ники'
    sheet["K2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['K2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['K3'].value = 11
    sheet["K3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['K3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['K1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    sheet.merge_cells('L1:M1')
    sheet['L1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['M1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['L1'].value = 'Иш ҳақи,сўм'
    sheet["L1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['L2'].value = 'механи-затор-ники'
    sheet['L2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["L2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['L3'].value = 12
    sheet['L3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["L3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['M2'].value = 'ишчи-ники'
    sheet['M2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["M2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['M3'].value = 13
    sheet['M3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["M3"].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('N1:O1')
    sheet['N1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['O1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['N1'].value = 'Ёнилғи сарфи, литр'
    sheet["N1"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['N2'].value = 'бир бир-ликка (Ҳисоб)'
    sheet['N2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["N2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['N3'].value = 14
    sheet['N3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["N3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['O2'].value = 'жами ҳаж-мига'
    sheet['O2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["O2"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['O3'].value = 15
    sheet['O3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["O3"].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('P1:Q1')
    sheet['P1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['Q1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['P1'].value = 'Техникадан фойдаланиш харажатлари, минг сўм'
    sheet["P1"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['P2'].value = 'Аморти-зация'
    sheet['P2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["P2"].alignment = Alignment(horizontal='justify', vertical='justify')
    # sheet['P3'].value = 16
    sheet['P3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["P3"].alignment = Alignment(horizontal='center', vertical='center')
    sheet['Q2'].value = 'Таъмир-лаш'
    sheet['Q2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["Q2"].alignment = Alignment(horizontal='justify', vertical='justify')
    # sheet['Q3'].value = 17
    sheet['Q3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet["Q3"].alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells('R1:R2')
    sheet['R1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['R2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['R1'].value = 'Бошқа (ўғит, уруғлик, кимё ва хок.) харажатлар, минг сўм'
    sheet["R1"].alignment = Alignment(horizontal='justify', vertical='justify')
    sheet['R3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    sheet.merge_cells('S1:S2')
    sheet['S1'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['S2'].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    sheet['S1'].value = 'Жами харажатлар, минг сўм'
    sheet["S1"].alignment = Alignment(horizontal='justify', vertical='justify')

    # sheet['Q3'].value = 17
    sheet['S3'].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    # sheet['B1'].style = "Good"
    # sheet['D1'].style = "Bad"
    # print(result)
    #

    full_path = os.path.join(current_app.root_path, 'saved_exl', f'{gis_area}-{ekin_id}')

    if not os.path.exists(full_path):
        os.makedirs(full_path)

    file_path = os.path.join(full_path, f'{gis_area}-{ekin_id}.xlsx')

    for row in ret_data:

        sheet.append({'a': row['name']})
        sheet[f'a{sheet.max_row}'].font = Font(bold=True, name='Arial', size=10, color='000000')
        for dict in row['arrangements']:
            sheet.append(dict)
            sheet.row_dimensions[sheet.max_row].height = 40
            for k in dict:
                # print(k)
                # key = get_key(dict, dict[k])

                sheet[f"{k}{sheet.max_row}"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
                sheet[f"{k}{sheet.max_row}"].alignment = Alignment(horizontal='justify', vertical='justify')

        sheet.append({'h': row['phase_overall']['days_of_shift'],
                      'i': row['phase_overall']['days_of_shift_human'],
                      'l': row['phase_overall']['work_machine'],
                      'm': row['phase_overall']['work_human'],
                      'o': row['phase_overall']['jami_xaj_miga']
                      })
        for row in sheet.columns:
            for value in row:

                if isinstance(value.value, float) or isinstance(value.value, int) or isinstance(value.value, str):
                    if len(str(value.value)) <= 8:
                        a = get_column_letter(value.column)

                        # print(f"{a}{value.row}-{value.value}")
                        sheet[f"{a}{value.row}"].alignment = Alignment(horizontal='center', vertical='center')
        # print(row['phase_overall']['days_of_shift'])

    workbook.save(output)
    output.seek(0)

    # return jsonify({'msg':f"{file_path}"})
    return Response(output, mimetype="application/ms-excel",
                    headers={"Content-Disposition": f"attachment;filename={cadastor}_{ekin_id}.xlsx"})


def get_key(d, value):
    for k, v in d.items():
        if v == value:
            return k


# import numpy as np

#
# @new_agro.get('/get_start_and_end')
# def StartEnd():
#     cadastor = request.args.get('cadastor')
#     gis_area = request.args.get('gis_area')
#
#     cashed = redis.get(f'startendTime:{cadastor}{gis_area}')
#
#     if cashed:
#         print('cashed')
#         data = json.loads(cashed)
#         return jsonify(data)
#
#     all_phases = db.session.query(Phase).join(Arrangement, Arrangement.phase_id == Phase.id).order_by(Phase.id).all()
#
#     arrs = Arrangement_copy.query.filter(Arrangement_copy.cadastor == cadastor, Arrangement_copy.gis_area == gis_area)
#     # group_arrs = db.session.query(Arrangement_copy.contour_number, db.func.count()).filter(Arrangement_copy.cadastor == cadastor, Arrangement_copy.gis_area == gis_area).group_by(Arrangement_copy.contour_number).all()
#     #
#     # print(group_arrs)
#     #
#     # group_arrs_ls = []
#     #
#     # for x in group_arrs:
#     #     group_arrs_ls.append(x[1])
#     #
#
#     arrs_contour = []
#
#     for i in arrs:
#         arrs_contour.append(i.contour_number)
#
#     # try:
#
#     lst = []
#     for contour in set(arrs_contour):
#         for x in all_phases:
#             ret_format = {
#                 'id': x.id,
#                 'name': x.name,
#                 'contour': contour,
#                 'arrangements': [],
#
#             }
#
#             for i in Arrangement_copy.query.filter(
#                     Arrangement_copy.cadastor == cadastor, Arrangement_copy.gis_area == gis_area,
#                     Arrangement_copy.phase_id == x.id, Arrangement_copy.contour_number == contour).order_by(
#                 Arrangement_copy.index):
#                 ret_format['arrangements'].append(i.format())
#
#             lst.append(ret_format)
#             # filt = filter(set(arrs_contour), lst)
#             # print(filt)
#
#     # overall_arrs = len(lst) - len(group_arrs_ls)
#     splits = np.array_split(lst, len(set(arrs_contour)))
#     print(len(set(arrs_contour)))
#     times = []
#
#     for i in splits:
#         # print(splits)
#         just_list = list(i)
#         # print(just_list)
#         first_phase_start_time = just_list[0]['arrangements'][0]['start_time']
#         plant_id = just_list[0]['arrangements'][0]['plant_id']
#         first_phase_end_time = just_list[2]['arrangements'][0]['start_time']
#         other_phase_start_time = just_list[2]['arrangements'][0]['start_time']
#         other_phase_end_time = just_list[-1]['arrangements'][-1]['end_time']
#         first_phase_name = just_list[0]['name']
#         contour = just_list[0]['contour']
#
#         obj = {
#             'first_phase_start_time': first_phase_start_time.strftime("%Y.%m.%d") if first_phase_start_time else None,
#             'first_phase_end_time': first_phase_end_time.strftime("%Y.%m.%d") if first_phase_end_time else None,
#             'other_phase_start_time': other_phase_start_time.strftime("%Y.%m.%d") if other_phase_start_time else None,
#             'other_phase_end_time': other_phase_end_time.strftime("%Y.%m.%d") if other_phase_end_time else None,
#             'first_phase_name': first_phase_name,
#             'contour': contour,
#             'plant_id': plant_id,
#         }
#
#         times.append(obj)
#
#     # redis.set(f'startendTime:{cadastor}{gis_area}', json.dumps(times),
#     #           ex=600)
#
#     return jsonify(times)
#     # except Exception as e:
#     print(e)
#     return jsonify({'msg': str(e)})


@new_agro.get('/all_start_end')
def allStartEnd():
    contour_number = request.args.get('contour_number')
    gis_area = request.args.get('gis_area')
    cadastor = request.args.get('cadastor')

    cashed = redis.get(f"all_start_end: {contour_number}-{gis_area}-{cadastor}")

    if cashed:
        print('cashed')
        return json.loads(cashed)

    savedStartEnd = SavedArrTime.query.filter_by(gis_area=gis_area, cadastor_number=cadastor,
                                                 contour_number=contour_number).order_by(
        SavedArrTime.id).all()

    redis.set(f"all_start_end: {contour_number}-{gis_area}-{cadastor}", json.dumps(list(map(lambda i: i.format(), savedStartEnd))), ex=120)

    return list(map(lambda i: i.format(), savedStartEnd))


@new_agro.post('/create_land_passport')
def createLandPassport() -> dict:
    # jd1
    bir = request.form.get('bir')
    asosi = request.form.get('asosi')
    ayrisimon = request.form.get('ayrisimon')
    turt = request.form.get('turt')
    gis_area = request.form.get('gis_area')
    cadastor_num = request.form.get('cadastor_num')
    counter_number = request.form.get('contour_number')
    # jd2
    kiyaligi = request.form.get('kiyaligi')
    # jd3
    toshlok_darja = request.form.get('toshlok_daraja')
    # jd4
    guruh = request.form.get('guruh')
    length = request.form.get('length')

    suv_taminoti = request.form.get('suv_taminoti')
    shorlanganligi = request.form.get('shorlanganligi')
    ekish_sharoiti = request.form.get('ekish_sharoiti')
    tuprokning_name = request.form.get('tuprokning_name')
    mintaqa = request.form.get('mintaqa')

    yosh_yoki_kichik = request.form.get('yosh_yoki_kichik')
    yirik_daraht = request.form.get('yirik_daraht')

    passport_exist = EarthPasport.query.filter_by(cadastor_num=cadastor_num, counter_number=counter_number,
                                                  gis_area=gis_area).first()

    tusiclar_proc = (int(bir) * 4 + int(asosi) * 20 + int(ayrisimon) * 78 + int(turt) * 176 + int(
        yosh_yoki_kichik) * 1.56 + int(yirik_daraht) * 5.3) / float(gis_area) * 100

    if tusiclar_proc < 5:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='5 % гача').first()
    elif tusiclar_proc > 5 and tusiclar_proc < 10:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='5...10').first()
    elif tusiclar_proc > 10 and tusiclar_proc < 15:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='10...15').first()
    elif tusiclar_proc > 15 and tusiclar_proc < 20:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='15...20').first()
    elif tusiclar_proc > 20 and tusiclar_proc < 25:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='20...25').first()
    elif tusiclar_proc > 25 and tusiclar_proc < 30:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='25...30').first()
    else:
        jd1 = JadvalOne.query.filter_by(tusiclar_egalagan='30...35').first()

    jd2 = JadvalTwo.query.filter_by(kiyaligi=kiyaligi).first()

    jd3 = JadvalThree.query.filter_by(toshlok_darja=toshlok_darja).first()

    if float(length) < 200:
        jd4 = JadvalFour.query.filter_by(guruh=guruh, length='<200').first()
    elif float(length) > 200 and float(length) < 400:
        jd4 = JadvalFour.query.filter_by(guruh=guruh, length='200-400').first()
    elif float(length) > 400 and float(length) < 600:
        jd4 = JadvalFour.query.filter_by(guruh=guruh, length='400-600').first()
    elif float(length) > 600 and float(length) < 1000:
        jd4 = JadvalFour.query.filter_by(guruh=guruh, length='600-1000').first()
    else:
        jd4 = JadvalFour.query.filter_by(guruh=guruh, length='>1000').first()

    # if tuprokning_name == int(1):
    #     print(tuprokning_name)
    #     tuproq = Tuproq.query.get(1)
    # elif tuprokning_name == int(2):
    #     tuproq = Tuproq.query.get(2)
    # elif tuprokning_name == 3:
    #     tuproq = Tuproq.query.get(3)
    # else:
    #     tuproq = Tuproq.query.get(4)

    tuproq = Tuproq.query.filter_by(tuprokning_name=tuprokning_name).first()

    if passport_exist:
        passport_exist.cadastor_num = cadastor_num
        passport_exist.counter_number = counter_number
        passport_exist.gis_area = gis_area
        passport_exist.one_jadval = jd1.id
        passport_exist.two_jadval = jd2.id
        passport_exist.three_jadval = jd3.id
        passport_exist.four_jadval = jd4.id
        passport_exist.bir = bir
        passport_exist.asosi = asosi
        passport_exist.ayrisimon = ayrisimon
        passport_exist.turt = turt
        passport_exist.tuproq_id = tuproq.id
        passport_exist.suv_taminoti = suv_taminoti
        passport_exist.shorlanganligi = shorlanganligi
        passport_exist.ekish_sharoiti = ekish_sharoiti
        passport_exist.mintaqa = mintaqa
        passport_exist.yosh_yoki_kichik = yosh_yoki_kichik
        passport_exist.yirik_daraht = yirik_daraht

        db.session.commit()

        passport_obj = {
            'kiyaligi': passport_exist.jadvaltwo.kiyaligi,
            'bir': passport_exist.bir,
            'asosi': passport_exist.asosi,
            'ayrisimon': passport_exist.ayrisimon,
            'turt': passport_exist.turt,
            'toshlok_darja': passport_exist.jadvalthree.toshlok_darja,
            'guruh': passport_exist.jadvalfour.guruh,
            'suv_taminoti': passport_exist.suv_taminoti,
            'shorlanganligi': passport_exist.shorlanganligi,
            'ekish_sharoiti': passport_exist.ekish_sharoiti,
            'tuproqning_name': passport_exist.tuproq_passport.tuprokning_name,
            'tuproqning_kh': passport_exist.tuproq_passport.tuprokning_kh,
            'mintaqa': passport_exist.mintaqa,
            'yosh_yoki_kichik': passport_exist.yosh_yoki_kichik,
            'yirik_daraht': passport_exist.yirik_daraht,
        }

        return jsonify(passport_obj)

    passport = EarthPasport(cadastor_num=cadastor_num, gis_area=gis_area, one_jadval=jd1.id,
                            two_jadval=jd2.id, three_jadval=jd3.id, four_jadval=jd4.id, counter_number=counter_number,
                            bir=bir,
                            asosi=asosi, ayrisimon=ayrisimon, turt=turt, tuproq_id=tuproq.id,
                            suv_taminoti=suv_taminoti, shorlanganligi=shorlanganligi, ekish_sharoiti=ekish_sharoiti,
                            yosh_yoki_kichik=yosh_yoki_kichik, yirik_daraht=yirik_daraht
                            )

    db.session.add(passport)
    db.session.commit()

    return jsonify(success)


@new_agro.get('/passport_values')
def passportValues() -> dict:
    gis_area = request.args.get('gis_area')
    cadastor = request.args.get('cadastor')
    contour_number = request.args.get('contour_number')
    try:
        passport = EarthPasport.query.filter_by(cadastor_num=cadastor, gis_area=gis_area,
                                                counter_number=contour_number).first()

        # er_haydashdan = passport.jadvalfour.er_hayd * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        # zanjirli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.zanjir
        # boshka_ishlarida = passport.jadvalfour.boshka_ish * passport.jadvalthree.ish_umum * passport.jadvalone.gildir_va_zanjir
        # gildirakli_tractor_bulsa = passport.jadvalthree.enligi * passport.jadvalone.gildir

        passport_obj = {
            'kiyaligi': passport.jadvaltwo.kiyaligi,
            'bir': passport.bir,
            'asosi': passport.asosi,
            'ayrisimon': passport.ayrisimon,
            'turt': passport.turt,
            'toshlok_darja': passport.jadvalthree.toshlok_darja,
            'guruh': passport.jadvalfour.guruh,
            'suv_taminoti': passport.suv_taminoti,
            'shorlanganligi': passport.shorlanganligi,
            'ekish_sharoiti': passport.ekish_sharoiti,
            'tuproqning_name': passport.tuproq_passport.tuprokning_name,
            'tuproqning_kh': passport.tuproq_passport.tuprokning_kh,
            'mintaqa': passport.mintaqa,
            'yosh_yoki_kichik': passport.yosh_yoki_kichik,
            'yirik_daraht': passport.yirik_daraht,
        }

        return jsonify(passport_obj)
    except Exception as e:
        print(e)
        return {}


import docx


@new_agro.get('/get_docx')
def GetDocx():
    doc = docx.Document('1-файл_иш_тури.docx')

    for paragraph in doc.paragraphs:
        if paragraph.text == '':
            pass
            # print(paragraph.text, 'ssss')
        else:
            # print(paragraph.text)
            digit = []
            text = []
            russ = []
            uzb = []
            for i in paragraph.text[0:5]:
                if i == '.':
                    digit.append(i)
                if i.isnumeric():
                    digit.append(i)
            for i in paragraph.text:
                s = paragraph.text

                if i == '(':
                    index = s.find(i)
                    # print(index)
                    if index > 0:
                        russ.append(paragraph.text[int(index)::])
                        uzb.append(paragraph.text[5:int(index)])

                if i == ' ':
                    text.append('_')
                if i.isnumeric():
                    pass
                else:
                    text.append(i)

            digit_str = ' '.join(digit).replace(' ', '')
            rus_str = ' '.join(russ).replace('.', '').replace('_', ' ').replace('(', '').replace(')', '')
            text_str = ' '.join(text).replace('.', '').replace(' ', '').replace('_', ' ').replace(f'({rus_str})', '')

            # # print(text)
            # print(digit_str)
            # print(str(text_str))
            # print(str(rus_str))

            extra_arr = Extara_Arrangements(
                name_ru=rus_str,
                name_uz=text_str,
                number=digit_str
            )
            extra_arr.save()

    return jsonify(success)


@new_agro.get('/get_extra_arrs')
def getExtraArrs():
    extra_arrs = Extara_Arrangements.query.all()
    cashed = redis.get('get_extra_arrs')

    if cashed:
        print('cashed')
        json.loads(cashed)
    result = list(map(lambda x: x.format(), extra_arrs))
    redis.set('get_extra_arrs', json.dumps(result), ex=600)

    return jsonify(result)


#
# @new_agro.post("/add")
# def start_add() -> dict:
#     a: int = request.form.get("a", type=int)
#     b: int = request.form.get("b", type=int)
#     result = add_together.delay(a, b)
#     print(result)
#     return {"result_id": result.id}
#
# #
# #
# # @celery.task
# # def add_together(a: int, b: int) -> int:
# #     print(a+b)
# #     return a + b
# @new_agro.get("/result/<id>")
# def task_result(id: str) -> dict:
#     # id: str = request.args.get('id')
#     result = add_together.AsyncResult(str(id))
#
#     # print(result.get())
#     # redis.flushdb()
#     return {
#         "ready": result.ready(),
#         "successful": result.successful(),
#         "value": result.result if result.ready() else None,
#     }
#
#
# app, celery = create_app()
#
#
# @celery.task()
# def add_together(a, b):
#     regions = Regions.query.all()
#     result = a + b
#     return [x.format() for x in regions]

@new_agro.post('/create_shablon_params')
def createShablon():
    ik_cols = list_of_cols_from(Shablon)

    ik = Shablon()
    try:
        for col in ik_cols:
            val = request.form.get(col)
            if val:
                setattr(ik, col, val)
                if col == 'ximicat':
                    # print(bool(int(request.form.get('ximicat'))))
                    ik.ximicat = bool(int(request.form.get('ximicat')))

        db.session.add(ik)
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.post('/update_shablon_params')
def updateShablon():
    id = request.form.get('id')
    ik_cols = list_of_cols_from(Shablon)

    shablon = Shablon.query.get(id)

    try:
        for col in ik_cols:
            val = request.form.get(col)
            if val:
                setattr(shablon, col, val)
                if col == 'ximicat':
                    # print(bool(int(request.form.get('ximicat'))))
                    shablon.ximicat = bool(int(request.form.get('ximicat')))
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.get('/delete_shablon_params')
def deleteShablon():
    id = request.args.get('id')

    shablon = Shablon.query.get(id)

    db.session.delete(shablon)
    db.session.commit()

    return jsonify(success)


@new_agro.get('/all_shablons_params')
def allShablons():
    shablons = Shablon.query.all()

    return list(map(lambda x: x.format(), shablons))


@new_agro.get('/get_shablon_params')
def getShablon():
    id = request.args.get('id')

    shablon = Shablon.query.get(id)

    return jsonify(shablon.format())


@new_agro.post('/set_shablon')
def set_shablon():
    ik_cols = list_of_cols_from(Shablon_Arrange)

    ik = Shablon_Arrange()
    try:
        for col in ik_cols:
            val = request.form.get(col)
            if val:
                setattr(ik, col, val)
        db.session.add(ik)
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.post('/update_shablon')
def update_shablon() -> Shablon_Arrange:
    id = request.form.get('id')
    ik_cols = list_of_cols_from(Shablon_Arrange)

    shablon = Shablon_Arrange.query.get(id)

    try:
        for col in ik_cols:
            val = request.form.get(col)
            if val:
                setattr(shablon, col, val)
        db.session.commit()
        return jsonify(success)
    except Exception as e:
        print(e)
        return jsonify({"error in col ": str(e)}), 445


@new_agro.get('/delete_shablon')
def delete_Shablon():
    id = request.args.get('id')

    shablon = Shablon_Arrange.query.get(id)

    db.session.delete(shablon)
    db.session.commit()

    return jsonify(success)


@new_agro.get('/all_shablons')
def all_Shablons():
    shablons = Shablon_Arrange.query.all()

    return list(map(lambda x: x.format(), shablons))


@new_agro.get('/get_shablon')
def get_Shablon() -> Shablon_Arrange:
    id = request.args.get('id')

    shablon = Shablon_Arrange.query.get(id)

    return jsonify(shablon.format())


@new_agro.get('/get_crops')
def getCrops() ->list:
    crops = Plan.query.all()

    return list(map(lambda x: x.crops(), crops))


@new_agro.post('/update_crop')
def updateCrop() -> dict:
    id = request.form.get('id')
    print(id)

    crop = Plan.query.get(id)

    crop.crop_name = request.form.get('crop_name')

    db.session.commit()

    return jsonify(success)

# @new_agro.get('/test_shablon')
# def test_shablon():
#     shablon = Shablon.query.get(1)
#
#     print(shablon.arrangements[0].arrangement.name)
#
#     return [x.arrangement.name for x in shablon.arrangements]


@new_agro.get('/spravochnik')
def spravochnik() -> list:
    jdone = JadvalOne.query.all()
    jdtwo = JadvalTwo.query.all()
    jdthree = JadvalThree.query.all()
    jdfour = JadvalFour.query.all()

    spravochnik_format = [
        {'jdone': [x.format() for x in jdone]},
        {'jdtwo': [x.format() for x in jdtwo]},
        {'jdthree': [x.format() for x in jdthree]},
        {'jdfour': [x.format() for x in jdfour]},
    ]

    return jsonify(spravochnik_format)

@new_agro.route('/shablon_ar')
def getShablonAr():
    option = request.args.get('option')
    if option == 'shablon_arrange':
        shablon = Shablon.query.all()
    else:
        shablon = Arrangement.query.all()
    return jsonify([x.format() for x in shablon])